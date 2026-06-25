#!/usr/bin/env python3
"""Download, edit, and upload a Realinsight model form Excel template.

This reference script intentionally uses the ri-agent CLI instead of calling
Core API directly. It downloads the workbook to disk, stages the edited file to
a temporary staged_file_id, then saves the template with that handle so the
final write call does not carry workbook bytes.

Example:
  python model_form_template_roundtrip.py \
    --model-form-id 665000000000000000000001 \
    --workdir /tmp/ri-model-form \
    --set-cell "Input!B2=Draft value" \
    --approved

Run once without --approved to download and edit locally. After reviewing the
workbook, rerun with --approved; the script reuses the existing workbook in
--workdir unless --force-download is supplied.

Requires:
  pip install openpyxl
  ri-agent auth login
"""

from __future__ import annotations

import argparse
import json
import pathlib
import subprocess
import sys
from typing import Any


def run_ri_agent(*args: str) -> dict[str, Any]:
    result = subprocess.run(
        ["ri-agent", *args],
        check=True,
        capture_output=True,
        text=True,
    )
    return json.loads(result.stdout)


def first_item(payload: dict[str, Any]) -> dict[str, Any]:
    items = payload.get("items") or []
    if not items:
        raise RuntimeError("ri-agent returned no items")
    return items[0]


def safe_file_name(file_name: str | None, fallback: str) -> str:
    if not file_name:
        return fallback
    safe_name = file_name.replace("\\", "/").split("/")[-1]
    return safe_name or fallback


def get_template_file_name(model_form_id: str, profile_args: list[str]) -> str:
    payload = run_ri_agent("download-model-form-template", model_form_id, *profile_args)
    file_name = first_item(payload).get("file_name")
    return safe_file_name(file_name, f"{model_form_id}.xlsx")


def get_conflict_token(model_form_id: str, profile_args: list[str]) -> str:
    payload = run_ri_agent("get-model-form", model_form_id, *profile_args)
    token = first_item(payload).get("conflict_token")
    if not token:
        raise RuntimeError("model form response did not include conflict_token")
    return token


def stage_template_file(model_form_id: str, template_path: pathlib.Path, profile_args: list[str]) -> str:
    payload = run_ri_agent(
        "stage-model-form-template",
        model_form_id,
        "--file-path",
        str(template_path),
        "--approved",
        *profile_args,
    )
    staged_file_id = first_item(payload).get("staged_file_id")
    if not staged_file_id:
        raise RuntimeError("template stage response did not include staged_file_id")
    return str(staged_file_id)


def parse_cell_assignment(value: str) -> tuple[str, str, str]:
    if "=" not in value or "!" not in value:
        raise ValueError("--set-cell must look like Sheet!A1=value")
    target, cell_value = value.split("=", 1)
    sheet_name, cell = target.split("!", 1)
    if not sheet_name or not cell:
        raise ValueError("--set-cell must include both sheet and cell")
    return sheet_name, cell, cell_value


def edit_workbook(path: pathlib.Path, assignments: list[str]) -> None:
    from openpyxl import load_workbook

    keep_vba = path.suffix.lower() == ".xlsm"
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise RuntimeError("This reference script edits .xlsx/.xlsm only.")

    workbook = load_workbook(path, keep_vba=keep_vba)
    for assignment in assignments:
        sheet_name, cell, value = parse_cell_assignment(assignment)
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"Sheet not found: {sheet_name}")
        workbook[sheet_name][cell] = value
    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--model-form-id", required=True)
    parser.add_argument("--workdir", required=True)
    parser.add_argument("--profile")
    parser.add_argument("--set-cell", action="append", default=[])
    parser.add_argument("--approved", action="store_true")
    parser.add_argument("--force-download", action="store_true")
    args = parser.parse_args()

    workdir = pathlib.Path(args.workdir).expanduser().resolve()
    workdir.mkdir(parents=True, exist_ok=True)
    profile_args = ["--profile", args.profile] if args.profile else []
    template_path = workdir / get_template_file_name(args.model_form_id, profile_args)

    should_download = args.force_download or not args.approved or not template_path.exists()
    if should_download:
        run_ri_agent(
            "download-model-form-template",
            args.model_form_id,
            "--output-path",
            str(template_path),
            *profile_args,
        )
    else:
        print(f"Using existing reviewed template at {template_path}")

    edit_workbook(template_path, args.set_cell)

    conflict_token = get_conflict_token(args.model_form_id, profile_args)
    if not args.approved:
        print(f"Edited template written to {template_path}")
        print("Review it, then rerun with --approved to upload the existing reviewed workbook.")
        return 0

    staged_file_id = stage_template_file(args.model_form_id, template_path, profile_args)
    payload = run_ri_agent(
        "upload-model-form-template",
        args.model_form_id,
        "--request-json",
        json.dumps({"staged_file_id": staged_file_id}),
        "--expected-conflict-token",
        conflict_token,
        "--approved",
        *profile_args,
    )
    print(json.dumps(payload, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
